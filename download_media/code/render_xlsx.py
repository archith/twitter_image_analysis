import pandas, os
import sys
import util
import url_table
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

color_dict_downloaded = {url_table.INSTAGRAM_DOM: colors.RED, \
                         url_table.YOUTUBE_DOM: colors.GREEN, \
                         url_table.IMGUR_DOM: colors.YELLOW, \
                         url_table.TINYPIC_DOM: colors.BLUE}

color_dict_not_downloaded = {url_table.INSTAGRAM_DOM: colors.DARKRED, \
                             url_table.YOUTUBE_DOM: colors.DARKGREEN, \
                             url_table.IMGUR_DOM: colors.DARKYELLOW, \
                             url_table.TINYPIC_DOM: colors.DARKBLUE}


def generate_dbg_xls(media_hdf5_file_path, xl_file_name='debug.xlsx'):
    _media_url_df = pandas.read_hdf(media_hdf5_file_path, '/url_group/readout')

    try:
        assert 'url_domain' in _media_url_df.columns
        assert 'media_downloaded' in _media_url_df.columns
    except:
        print ('Dataframe doesnt have expected columns!')
        raise
    _media_url_df.to_excel(xl_file_name)
    _num_rows = len(_media_url_df)
    _num_cols = len(_media_url_df.columns)
    _wb = openpyxl.load_workbook(xl_file_name)
    _sheet_names = _wb.get_sheet_names()
    _ws = _wb.get_sheet_by_name(_sheet_names[0])

    # pre process so that they all look uniform in OpenOffice
    for row in range(2, _num_rows + 2):
        for col in range(2, _num_cols + 2):
            cell = _ws.cell(column=col, row=row)
            cell.font = Font(color=colors.BLACK)

    _wb.save(xl_file_name)


def colorize_xl(xl_file_name, domain_string=url_table.INSTAGRAM_DOM):
    # color code row according to the status of their media files
    _df = pandas.read_excel(xl_file_name)
    try:
        assert 'url_domain' in _df.columns
        assert 'media_downloaded' in _df.columns
    except:
        print ('Dataframe doesnt have expected columns!')
        raise

    _downloaded_entries = _df['media_downloaded'] == True
    _media_entries = _df['url_domain'] == domain_string
    try:
        _dl_color = color_dict_downloaded[domain_string]
        _n_dl_color = color_dict_not_downloaded[domain_string]
    except KeyError:
        print ('In correct type of media domain!')
        raise
    # the +2 is to translate to excel sheet cords with start with 1 and have a row with headers
    _dl_media_entries = _df[((_downloaded_entries & _media_entries) == True)].index + 2
    _n_dl_media_entries = _df[((_downloaded_entries == False) & _media_entries) == True].index + 2
    _wb = openpyxl.load_workbook(xl_file_name)

    _num_rows = len(_df)
    _num_cols = len(_df.columns)

    _sheet_names = _wb.get_sheet_names()
    _ws = _wb.get_sheet_by_name(_sheet_names[0])
    # color code rows according to their media properties
    for row in range(2, _num_rows + 2):
        if row in _dl_media_entries:
            for col in range(2, _num_cols + 2):
                cell = _ws.cell(column=col, row=row)
                cell.font = Font(color=_dl_color)
        if row in _n_dl_media_entries:
            for col in range(2, _num_cols + 2):
                cell = _ws.cell(column=col, row=row)
                cell.font = Font(color=_n_dl_color)

    _wb.save(xl_file_name)


def render_dbg_xls(json_file, excel_file, config=util.Config()):
    media_hdf5_file_path, _, _, _ = util.get_media_file_paths(json_file)

    generate_dbg_xls(media_hdf5_file_path, excel_file)
    print('Generated excel sheet.')

    print('Colorizing excel sheet.')

    if config.download_instagram:
        colorize_xl(excel_file, domain_string=url_table.INSTAGRAM_DOM)
    if config.download_yt:
        colorize_xl(excel_file, domain_string=url_table.YOUTUBE_DOM)
    if config.download_imgur:
        colorize_xl(excel_file, domain_string=url_table.IMGUR_DOM)
    if config.download_tinypic:
        colorize_xl(excel_file, domain_string=url_table.TINYPIC_DOM)


def generate_dbg_csv(media_hdf5_file_path, csv_file):
    _media_url_df = pandas.read_hdf(media_hdf5_file_path, '/url_group/readout')
    _media_url_df.to_csv(csv_file)


def render_dbg_csv(json_file, csv_file):
    media_hdf5_file_path, _, _, _ = util.get_media_file_paths(json_file)
    generate_dbg_csv(media_hdf5_file_path, csv_file)
